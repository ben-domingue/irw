import csv
import openpyxl

SRC = "../automated_finding/downloads/cos101/Responses_COS101_dataset.xlsx"
OUT = "../automated_finding/output_noncore/cos101_2026_openended.csv"

wb = openpyxl.load_workbook(SRC, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
item_cols = header[1:]  # Q1..Q3, full question text embedded in the header

out = []
for row in rows[1:]:
    student_id = row[0]
    for i, text in enumerate(row[1:], start=1):
        if text and str(text).strip():
            out.append({"id": student_id, "item": f"Q{i}", "text": str(text).strip()})

with open(OUT, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=["id", "item", "text"])
    w.writeheader()
    w.writerows(out)

print(f"{len(out)} rows, {len({r['id'] for r in out})} students")
print("item text (for itemtext reference, not shipped in resp table):")
for i, q in enumerate(item_cols, start=1):
    print(f"  Q{i}: {q}")
