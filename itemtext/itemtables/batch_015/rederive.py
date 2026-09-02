#!/usr/bin/env python3
"""Re-derive batch_015's shipped item/option text from the cached Dataverse sources.

Written under issue #1774. The point is that each verify_<table>.R can diff the
shipped CSV against text rebuilt from the source rather than against a stored
copy of itself -- an `evidence` string cannot be re-run, a rebuild can.

Run from itemtext/:  python3 itemtables/batch_015/rederive.py
Writes itemtables/batch_015/rederived.json as {table: {item: {"text":..., "opts":{resp:label}}}}.
Sources are re-fetched from the Dataverse API if the cache is absent.
"""
import json, os, re, io, sys, urllib.request

CACHE = ".cache"
API = "https://dataverse.harvard.edu/api/access/datafile/%s"

def fetch(path, fid, original=False):
    if os.path.exists(path): return path
    os.makedirs(os.path.dirname(path), exist_ok=True)
    url = API % fid + ("?format=original" if original else "")
    req = urllib.request.Request(url, headers={"User-Agent": "curl/8.0"})  # WAF filters on UA
    with urllib.request.urlopen(req, timeout=120) as r, open(path, "wb") as f:
        f.write(r.read())
    return path

def xl_header(path, row=0):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return [("" if c is None else str(c)).strip() for c in rows[row]]

def pdf_text(path):
    from pypdf import PdfReader
    return "\n".join((p.extract_text() or "") for p in PdfReader(path).pages)

def docx_text(path):
    import zipfile, html as H
    x = zipfile.ZipFile(path).read("word/document.xml").decode("utf-8", "replace")
    x = re.sub(r"</w:p>", "\n", x)
    return re.sub(r"[ \t]+", " ", H.unescape(re.sub(r"<[^>]+>", "", x)))

out = {}

# ---- weber2026_name_knowledge: German items + English twins ------------------
d = CACHE + "/weber2026_name_knowledge"
en = {}
for r in xl_header(fetch(d + "/item_wording.xlsx", "14086851"), 0) and \
        __import__("openpyxl").load_workbook(d + "/item_wording.xlsx", read_only=True,
            data_only=True)[ "Item_wording_English" ].iter_rows(values_only=True):
    if r[0] and str(r[0]).strip().startswith("X"): en[str(r[0]).strip()] = str(r[1]).strip()
gt = re.sub(r"\s+", " ", pdf_text(fetch(d + "/german_questionnaire.pdf", "13406258")))
ANCH = "Stimme überhaupt nicht zu Stimme nicht zu Teils/Teils Stimme zu Stimme voll und ganz zu"
seg = gt[gt.find(ANCH) + len(ANCH):]
de = [re.sub(r"\s+", " ", x).strip() for x in re.findall(r"([^☐]{10,300}?)\s*(?:☐\s*){5}", seg)]
de = [re.sub(r"^.*?Stimme voll und ganz zu\s*", "", x).strip()
      if "Teils/Teils" in x and "Stimme voll und ganz zu" in x else x for x in de]
GOPT = ["Stimme überhaupt nicht zu","Stimme nicht zu","Teils/Teils","Stimme zu","Stimme voll und ganz zu"]
out["weber2026_name_knowledge"] = {
    c: {"text": de[i], "translated": en[c], "opts": {str(k+1): GOPT[k] for k in range(5)}}
    for i, c in enumerate(sorted(en, key=lambda s: int(s[1:]))) if i < len(de)}

# ---- boudreau_2022_police_reform -------------------------------------------
d = CACHE + "/boudreau_2022_police_reform"
t = re.sub(r"\s+", " ", docx_text(fetch(d + "/codebook.docx", "6381236")))
res = {}
for c in ["certif", "carotid", "review", "interv"]:
    i = t.find(c + "All things considered")
    seg = t[i + len(c):]
    seg = seg[:seg.find(".No answer/missing")]
    m = re.match(r"(All things considered.*?\?)\s*(.*)$", seg)
    parts = [p.strip() for p in re.split(
        r"(?=Strongly support|Somewhat support|Somewhat oppose|Strongly oppose|Don’t know)",
        m.group(2)) if p.strip()]
    res[c] = {"text": m.group(1).strip(), "opts": {str(k+1): p for k, p in enumerate(parts)}}
out["boudreau_2022_police_reform"] = res

# ---- kushnir2017_anrt ------------------------------------------------------
d = CACHE + "/kushnir2017_anrt"
hdr = xl_header(fetch(d + "/data.xlsx", "3059090"))
KOPT = ["Strongly disagree","Do NOT really agree","More or less agree","Generally agree","Fully agree"]
out["kushnir2017_anrt"] = {h: {"text": re.sub(r"^\(ANRT_\d+\)\s*", "", h).strip(),
                               "opts": {str(k+1): KOPT[k] for k in range(5)}}
                           for h in hdr if h.startswith("(ANRT_")}

# ---- karajko x3 -----------------------------------------------------------
d = CACHE + "/karajko2025_ai_governance"
hdr = xl_header(fetch(d + "/ESM_1.xlsx", "11076712"))
KAR = {"karajko2025_ai_benefit": ("7", ["Uopšte ne","Ne previše","Onako","Donekle","Mnogo"]),
       "karajko2025_ai_risk": ("8", ["Snažno ne odobravam","Ne odobravam","Neutralan","Odobravam","Snažno odobravam"]),
       "karajko2025_ai_governance": ("12", ["Uopšte nije važno","Malo važno","Umjereno važno","Važno","Vrlo važno"]),
       "karajko2025_ai_trust": ("14", ["Uopšte ne","Ne previše","Onako","Donekle","Mnogo"])}
for tbl, (blk, opts) in KAR.items():
    res = {}
    for h in hdr:
        m = re.match(r"^P%s_(\d+)\." % blk, h, re.I)
        if m:
            res["p%s_%s" % (blk, m.group(1))] = {
                "text": re.sub(r"^P\d+_\d+\.\s*", "", re.sub(r"\s+", " ", h)).strip(),
                "opts": {str(k+1): opts[k] for k in range(5)}}
    out[tbl] = res

# ---- balmas x2 ------------------------------------------------------------
d = CACHE + "/balmas2018_leader_personality"
t = re.sub(r"[ \t]+", " ", pdf_text(fetch(d + "/Codebooks.pdf", "3109073")))
ent = {}
for m in re.finditer(r"(?m)^\s*\d+\.\s+(\S+)\s+(.*?)(?=^\s*\d+\.\s+\S+\s|\Z)", t, re.S):
    ent.setdefault(m.group(1), re.sub(r"\s+", " ", m.group(2)).strip())
VAL = re.compile(r'\s*(\d+)\s+"([^"]*)"')
def split(b):
    m = VAL.search(b)
    if not m: return re.sub(r"\s+", " ", b).strip(), {}
    return re.sub(r"\s+", " ", b[:m.start()]).strip(), {a: c for a, c in VAL.findall(b[m.start():])}
for tbl, codes in [("balmas2018_leader_personality",
                    ["c1","c2","c3","c4","c5","c8","c10","c11","c13","c14","c15","c16","c18","c26"]),
                   ("balmas2018_leader_attitudes", ["a3","a4","a5"])]:
    res = {}
    for c in codes:
        lab, vals = split(ent[c])
        res[c] = {"text": lab, "opts": {k: v for k, v in vals.items() if k in ("1","5")}}
    out[tbl] = res

# ---- lee2024_relative_clause: recover the row-2 header ---------------------
d = CACHE + "/lee2024_relative_clause"
row2 = xl_header(fetch(d + "/highschool.xlsx", "10677280", original=True), row=2)
out["lee2024_relative_clause"] = {"__header_row2__": row2}

# ---- Adherence_Zissette_2018_SDB: inline survey numbering ------------------
d = CACHE + "/Adherence_Zissette_2018_SDB"
t = re.sub(r"\s+", " ", re.sub(r"[ \t]+", " ", pdf_text(fetch(d + "/survey.pdf", "3137673"))))
num = {}
for m in re.finditer(r'(?<![\d.])(\d{1,3})\s+([A-Z‘“\'"][^\d]{8,300}?)\s*1 2 3 4 5 6', t):
    num.setdefault(int(m.group(1)), re.sub(r"\s+", " ", m.group(2)).strip())
ZOPT = ["Disagree a Lot","Disagree Somewhat","Disagree a Little","Agree a Little","Agree Somewhat","Agree a Lot"]
WANT = list(range(25,33)) + list(range(56,64)) + list(range(84,94))
out["Adherence_Zissette_2018_SDB"] = {
    "q_s_%03d_SDB" % n: {"text": num[n], "opts": {str(k+1): ZOPT[k] for k in range(6)}}
    for n in WANT if n in num}

p = "itemtables/batch_015/rederived.json"
json.dump(out, io.open(p, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("re-derived %d tables -> %s" % (len(out), p))
for k, v in out.items(): print("   %-32s %d entries" % (k, len(v)))
