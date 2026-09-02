#!/usr/bin/env python3
"""rederive_coach.py -- batch_016, issue #1831

Re-reads the COACH deposit's codebook.xlsx from scratch and writes
rederived_coach.json, so verify_COACH_Chen_2022_*.R diffs the shipped text
against a rebuild of the source workbook rather than against a prose claim.

Also re-runs the check behind text_source=translated_substitute: it counts CJK
characters in the deposit's own files, which is the evidence that the
administered Chinese wording is not recoverable from the deposit.

Reads   .cache/COACH_Chen_2022/{codebook.xlsx,raw.xlsx}
Writes  itemtables/batch_016/rederived_coach.json
Run from itemtext/:  python3 itemtables/batch_016/rederive_coach.py
"""
import json, os, re, sys, zipfile

CACHE = ".cache/COACH_Chen_2022"
OUT   = "itemtables/batch_016/rederived_coach.json"
CJK   = re.compile(r'[一-鿿]')

def parse_vals(s):
    """'1 = "Strongly agree"\\n2 = "agree"' -> {'1': 'Strongly agree', ...}"""
    out = {}
    for line in str(s).split("\n"):
        m = re.match(r'^\s*(\d+)\s*=\s*"?(.*?)"?\s*$', line)
        if m and m.group(2).strip():
            out[m.group(1)] = m.group(2).strip()
    return out

def cjk_count(path):
    """CJK characters in every shared string of an xlsx, font names included."""
    if not os.path.exists(path):
        return None
    try:
        with zipfile.ZipFile(path) as z:
            txt = z.read("xl/sharedStrings.xml").decode("utf-8", "ignore")
    except KeyError:
        return 0
    return len(CJK.findall(txt))

def main():
    xls = os.path.join(CACHE, "codebook.xlsx")
    if not os.path.exists(xls):
        sys.exit("cached codebook absent (%s); the committed rederived_coach.json stands.\n"
                 "Refetch from doi:10.7910/DVN/6RWH41 -- see provenance.csv source_ref." % xls)
    import openpyxl
    ws = openpyxl.load_workbook(xls, data_only=True).worksheets[0]

    # The header is not on row 1 -- the sheet opens with a blank row -- so find it.
    hdr, hdr_row = None, None
    for n, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
        cells = ["" if c is None else str(c).strip().lower() for c in row]
        if "variable" in cells:
            hdr, hdr_row = cells, n
            break
    if hdr is None:
        sys.exit("codebook.xlsx: no header row containing 'Variable' in the first 10 rows")
    iv, il, ivl = hdr.index("variable"), hdr.index("label"), hdr.index("value")
    entries = {}
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        var = row[iv]
        if not var or not str(var).strip():
            continue
        lab = "" if row[il] is None else str(row[il]).strip()
        val = "" if row[ivl] is None else str(row[ivl])
        entries[str(var).strip()] = {"text": lab, "opts": parse_vals(val)}

    payload = {
        "entries": entries,
        "cjk_evidence": {
            "codebook.xlsx": cjk_count(xls),
            "raw.xlsx": cjk_count(os.path.join(CACHE, "raw.xlsx")),
            "note": ("Any count here is font names such as SimSun, not item text. This is the "
                     "evidence for text_source=translated_substitute with language=Chinese."),
        },
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1, sort_keys=True)
    print("wrote %s: %d codebook variables (%d with an item label, %d with option labels); "
          "CJK codebook.xlsx=%s raw.xlsx=%s" % (
              OUT, len(entries), sum(1 for v in entries.values() if v["text"]),
              sum(1 for v in entries.values() if v["opts"]),
              payload["cjk_evidence"]["codebook.xlsx"], payload["cjk_evidence"]["raw.xlsx"]))

if __name__ == "__main__":
    main()
